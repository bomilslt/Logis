"""
Service d'Export - Génération PDF et Excel
==========================================

Génère des documents PDF (factures, étiquettes, rapports)
et des exports Excel pour les données.
"""

import io
import logging
from datetime import datetime
from typing import List, Optional
from dataclasses import dataclass

logger = logging.getLogger(__name__)

# Import conditionnel des dépendances
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab non installé - pip install reportlab")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl non installé - pip install openpyxl")


@dataclass
class ExportResult:
    """Résultat d'un export"""
    success: bool
    data: Optional[bytes] = None
    filename: Optional[str] = None
    content_type: Optional[str] = None
    error: Optional[str] = None


class PDFGenerator:
    """Générateur de documents PDF"""
    
    def __init__(self, tenant_name: str = "Express Cargo"):
        self.tenant_name = tenant_name
        self.styles = getSampleStyleSheet() if REPORTLAB_AVAILABLE else None
    
    def _get_primary_color(self, tenant_info: dict) -> colors.Color:
        """Récupère la couleur primaire depuis les settings"""
        color_hex = tenant_info.get('primary_color', '#2563eb') if tenant_info else '#2563eb'
        try:
            return colors.HexColor(color_hex)
        except:
            return colors.HexColor('#2563eb')
    
    def _add_logo(self, elements, tenant_info: dict):
        """Ajoute le logo avec aspect ratio préservé"""
        if not tenant_info:
            return
        
        show_logo = tenant_info.get('show_logo', True)
        logo_data = tenant_info.get('logo', '')
        
        if not show_logo or not logo_data or not logo_data.startswith('data:image'):
            return
        
        try:
            import base64
            from reportlab.lib.utils import ImageReader
            
            # Valider le format
            if ',' not in logo_data:
                logger.warning("Invalid logo format - missing comma separator")
                return
            
            header, encoded = logo_data.split(',', 1)
            
            # Valider le type d'image
            if not header.startswith('data:image/'):
                logger.warning("Invalid image type")
                return
            
            # Décoder et valider la taille
            try:
                logo_bytes = base64.b64decode(encoded)
            except Exception as e:
                logger.warning(f"Failed to decode base64 logo: {e}")
                return
            
            # Limiter la taille (max 500KB)
            if len(logo_bytes) > 500 * 1024:
                logger.warning("Logo too large (>500KB)")
                return
            
            logo_buffer = io.BytesIO(logo_bytes)
            
            # Lire les dimensions réelles de l'image
            try:
                img_reader = ImageReader(logo_buffer)
                img_width, img_height = img_reader.getSize()
            except Exception as e:
                logger.warning(f"Cannot read image dimensions: {e}")
                return
            
            if img_width <= 0 or img_height <= 0:
                logger.warning("Invalid image dimensions")
                return
            
            # Calculer les dimensions avec aspect ratio préservé
            aspect_ratio = img_height / float(img_width)
            max_width = 4 * cm
            max_height = 2 * cm
            
            # Adapter aux limites en préservant l'aspect ratio
            if aspect_ratio > max_height / max_width:
                # Image plus haute que large - limiter par la hauteur
                logo_height = max_height
                logo_width = logo_height / aspect_ratio
            else:
                # Image plus large que haute - limiter par la largeur
                logo_width = max_width
                logo_height = logo_width * aspect_ratio
            
            # Créer l'image avec les bonnes dimensions
            logo_buffer.seek(0)  # Reset buffer position
            img = Image(logo_buffer, width=logo_width, height=logo_height)
            img.hAlign = 'LEFT'
            img.vAlign = 'TOP'
            
            elements.append(img)
            elements.append(Spacer(1, 0.3*cm))
            
            logger.info(f"Logo added successfully: {logo_width/cm:.1f}x{logo_height/cm:.1f}cm")
            
        except Exception as e:
            logger.warning(f"Failed to add logo: {e}")
            # Continue without logo rather than failing
    
    def _add_custom_header(self, elements, tenant_info: dict):
        """Ajoute le header personnalisé si défini"""
        if not tenant_info:
            return
        
        header_text = tenant_info.get('header', '')
        if header_text:
            header_style = ParagraphStyle(
                'CustomHeader',
                parent=self.styles['Normal'],
                fontSize=9,
                textColor=colors.grey,
                alignment=1  # Centré
            )
            elements.append(Paragraph(header_text, header_style))
            elements.append(Spacer(1, 0.3*cm))
    
    def _add_custom_footer(self, elements, tenant_info: dict):
        """Ajoute le footer personnalisé si défini"""
        if not tenant_info:
            return
        
        footer_text = tenant_info.get('footer', '') or tenant_info.get('export_footer', '')
        if footer_text:
            footer_style = ParagraphStyle(
                'CustomFooter',
                parent=self.styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=1
            )
            elements.append(Spacer(1, 1*cm))
            elements.append(Paragraph(footer_text, footer_style))
    
    def generate_invoice_pdf(self, invoice: dict, tenant_info: dict = None) -> ExportResult:
        """
        Génère un PDF de facture
        
        Args:
            invoice: Données de la facture (dict)
            tenant_info: Infos du tenant (nom, adresse, logo, header, footer, primary_color)
        
        Returns:
            ExportResult avec le PDF en bytes
        """
        if not REPORTLAB_AVAILABLE:
            return ExportResult(success=False, error="reportlab non installé")
        
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                   rightMargin=2*cm, leftMargin=2*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)
            
            elements = []
            styles = self.styles
            primary_color = self._get_primary_color(tenant_info)
            
            # Logo
            self._add_logo(elements, tenant_info)
            
            # Header personnalisé
            self._add_custom_header(elements, tenant_info)
            
            # Style personnalisé pour le titre avec couleur primaire
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                textColor=primary_color
            )
            
            # En-tête entreprise
            tenant_name = tenant_info.get('name', self.tenant_name) if tenant_info else self.tenant_name
            elements.append(Paragraph(tenant_name, title_style))
            
            if tenant_info:
                if tenant_info.get('address'):
                    elements.append(Paragraph(tenant_info['address'], styles['Normal']))
                if tenant_info.get('phone'):
                    elements.append(Paragraph(f"Tél: {tenant_info['phone']}", styles['Normal']))
                if tenant_info.get('email'):
                    elements.append(Paragraph(f"Email: {tenant_info['email']}", styles['Normal']))
            
            elements.append(Spacer(1, 1*cm))
            
            # Titre facture
            elements.append(Paragraph(f"FACTURE N° {invoice.get('invoice_number', 'N/A')}", 
                                     styles['Heading2']))
            elements.append(Spacer(1, 0.5*cm))
            
            # Infos facture
            info_data = [
                ['Date d\'émission:', invoice.get('issue_date', 'N/A')],
                ['Date d\'échéance:', invoice.get('due_date', 'N/A')],
                ['Statut:', invoice.get('status', 'N/A').upper()],
            ]
            
            info_table = Table(info_data, colWidths=[4*cm, 6*cm])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.5*cm))
            
            # Client
            elements.append(Paragraph("FACTURER À:", styles['Heading3']))
            elements.append(Paragraph(invoice.get('client_name', 'N/A'), styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
            
            # Description et montant
            elements.append(Paragraph("DÉTAILS:", styles['Heading3']))
            
            detail_data = [
                ['Description', 'Montant'],
                [invoice.get('description', 'N/A'), 
                 f"{invoice.get('amount', 0):,.0f} {invoice.get('currency', 'XAF')}"]
            ]
            
            if invoice.get('package_tracking'):
                detail_data[1][0] += f"\nColis: {invoice['package_tracking']}"
            
            detail_table = Table(detail_data, colWidths=[12*cm, 4*cm])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), primary_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(detail_table)
            elements.append(Spacer(1, 0.5*cm))
            
            # Total
            total_data = [
                ['TOTAL:', f"{invoice.get('amount', 0):,.0f} {invoice.get('currency', 'XAF')}"]
            ]
            total_table = Table(total_data, colWidths=[12*cm, 4*cm])
            total_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 14),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f0f0')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(total_table)
            
            # Notes
            if invoice.get('notes'):
                elements.append(Spacer(1, 1*cm))
                elements.append(Paragraph("Notes:", styles['Heading4']))
                elements.append(Paragraph(invoice['notes'], styles['Normal']))
            
            # Footer personnalisé
            self._add_custom_footer(elements, tenant_info)
            
            # Pied de page par défaut
            elements.append(Spacer(1, 1*cm))
            elements.append(Paragraph("Merci pour votre confiance!", 
                                     ParagraphStyle('Footer', parent=styles['Normal'],
                                                   alignment=1, textColor=colors.grey)))
            
            doc.build(elements)
            
            pdf_data = buffer.getvalue()
            buffer.close()
            
            filename = f"facture_{invoice.get('invoice_number', 'unknown').replace('/', '-')}.pdf"
            
            return ExportResult(
                success=True,
                data=pdf_data,
                filename=filename,
                content_type='application/pdf'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération PDF facture: {str(e)}")
            return ExportResult(success=False, error=str(e))
    
    def generate_package_label_pdf(self, package: dict, tenant_info: dict = None) -> ExportResult:
        """
        Génère une étiquette de colis en PDF
        
        Args:
            package: Données du colis
            tenant_info: Infos du tenant (logo, primary_color)
        
        Returns:
            ExportResult avec le PDF
        """
        if not REPORTLAB_AVAILABLE:
            return ExportResult(success=False, error="reportlab non installé")
        
        try:
            buffer = io.BytesIO()
            # Étiquette format 10x15 cm
            page_width = 10 * cm
            page_height = 15 * cm
            
            c = canvas.Canvas(buffer, pagesize=(page_width, page_height))
            
            # Couleur primaire pour la bordure
            primary_color = self._get_primary_color(tenant_info)
            
            # Bordure avec couleur primaire
            c.setStrokeColor(primary_color)
            c.setLineWidth(2)
            c.rect(5*mm, 5*mm, page_width - 10*mm, page_height - 10*mm)
            
            # Nom du tenant avec couleur primaire
            tenant_name = tenant_info.get('name', self.tenant_name) if tenant_info else self.tenant_name
            c.setFillColor(primary_color)
            c.setFont("Helvetica-Bold", 14)
            c.drawCentredString(page_width/2, page_height - 15*mm, tenant_name)
            
            # Reset couleur pour le reste
            c.setFillColor(colors.black)
            
            # Tracking number (gros)
            c.setFont("Helvetica-Bold", 18)
            c.drawCentredString(page_width/2, page_height - 30*mm, 
                              package.get('tracking_number', 'N/A'))
            
            # Ligne de séparation avec couleur primaire
            c.setStrokeColor(primary_color)
            c.setLineWidth(1)
            c.line(10*mm, page_height - 38*mm, page_width - 10*mm, page_height - 38*mm)
            
            # Destinataire
            y_pos = page_height - 50*mm
            c.setFont("Helvetica-Bold", 10)
            c.drawString(10*mm, y_pos, "DESTINATAIRE:")
            
            c.setFont("Helvetica", 10)
            recipient = package.get('recipient', {})
            y_pos -= 5*mm
            c.drawString(10*mm, y_pos, recipient.get('name', 'N/A'))
            y_pos -= 4*mm
            c.drawString(10*mm, y_pos, recipient.get('phone', ''))
            
            # Destination
            y_pos -= 8*mm
            c.setFont("Helvetica-Bold", 10)
            c.drawString(10*mm, y_pos, "DESTINATION:")
            
            c.setFont("Helvetica", 10)
            dest = package.get('destination', {})
            y_pos -= 5*mm
            c.drawString(10*mm, y_pos, dest.get('warehouse', dest.get('city', 'N/A')))
            y_pos -= 4*mm
            c.drawString(10*mm, y_pos, dest.get('country', ''))
            
            # Infos colis
            y_pos -= 10*mm
            c.setFont("Helvetica-Bold", 9)
            c.drawString(10*mm, y_pos, f"Poids: {package.get('weight', 'N/A')} kg")
            c.drawString(page_width/2, y_pos, f"Qté: {package.get('quantity', 1)}")
            
            y_pos -= 5*mm
            c.drawString(10*mm, y_pos, f"Mode: {package.get('transport_mode', 'N/A')}")
            
            # Date
            y_pos -= 10*mm
            c.setFont("Helvetica", 8)
            c.drawString(10*mm, y_pos, f"Créé le: {package.get('created_at', 'N/A')[:10]}")
            
            c.save()
            
            pdf_data = buffer.getvalue()
            buffer.close()
            
            filename = f"etiquette_{package.get('tracking_number', 'unknown')}.pdf"
            
            return ExportResult(
                success=True,
                data=pdf_data,
                filename=filename,
                content_type='application/pdf'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération étiquette: {str(e)}")
            return ExportResult(success=False, error=str(e))
    
    def generate_payment_receipt(self, payment: dict, tenant_info: dict = None) -> ExportResult:
        """
        Génère un reçu de paiement en PDF
        
        Args:
            payment: Données du paiement
            tenant_info: Infos du tenant (logo, header, footer, primary_color)
        
        Returns:
            ExportResult avec le PDF
        """
        if not REPORTLAB_AVAILABLE:
            return ExportResult(success=False, error="reportlab non installé")
        
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4,
                                   rightMargin=2*cm, leftMargin=2*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)
            
            elements = []
            styles = self.styles
            primary_color = self._get_primary_color(tenant_info)
            
            # === HEADER SECTION ===
            # Logo et nom entreprise côte à côte avec layout propre
            tenant_name = tenant_info.get('name', self.tenant_name) if tenant_info else self.tenant_name
            
            # Créer le header avec logo à gauche et infos à droite
            if tenant_info and tenant_info.get('logo') and tenant_info.get('logo').startswith('data:image'):
                try:
                    import base64
                    from reportlab.lib.utils import ImageReader
                    
                    logo_data = tenant_info['logo']
                    if ',' not in logo_data:
                        raise ValueError("Invalid logo format")
                    
                    header, encoded = logo_data.split(',', 1)
                    logo_bytes = base64.b64decode(encoded)
                    
                    # Limiter la taille
                    if len(logo_bytes) > 500 * 1024:
                        raise ValueError("Logo too large")
                    
                    logo_buffer = io.BytesIO(logo_bytes)
                    
                    # Lire dimensions et calculer aspect ratio
                    img_reader = ImageReader(logo_buffer)
                    img_width, img_height = img_reader.getSize()
                    aspect_ratio = img_height / float(img_width) if img_width > 0 else 1
                    
                    # Calculer dimensions du logo pour le header
                    max_logo_width = 4 * cm
                    max_logo_height = 2 * cm
                    
                    if aspect_ratio > max_logo_height / max_logo_width:
                        logo_height = max_logo_height
                        logo_width = logo_height / aspect_ratio
                    else:
                        logo_width = max_logo_width
                        logo_height = logo_width * aspect_ratio
                    
                    # Créer l'image du logo
                    logo_buffer.seek(0)
                    logo_img = Image(logo_buffer, width=logo_width, height=logo_height)
                    
                    # Créer les informations entreprise dans un tableau vertical
                    company_info_data = []
                    
                    # Nom de l'entreprise
                    company_name_para = Paragraph(
                        f"<b>{tenant_name}</b>",
                        ParagraphStyle(
                            'CompanyName',
                            parent=styles['Normal'],
                            fontSize=14,
                            textColor=primary_color,
                            spaceAfter=4
                        )
                    )
                    company_info_data.append([company_name_para])
                    
                    # Adresse
                    if tenant_info.get('address'):
                        address_para = Paragraph(
                            tenant_info['address'],
                            ParagraphStyle(
                                'CompanyAddr',
                                parent=styles['Normal'],
                                fontSize=9,
                                spaceAfter=2
                            )
                        )
                        company_info_data.append([address_para])
                    
                    # Contact (téléphone et email)
                    contact_parts = []
                    if tenant_info.get('phone'):
                        contact_parts.append(f"Tél: {tenant_info['phone']}")
                    if tenant_info.get('email'):
                        contact_parts.append(f"Email: {tenant_info['email']}")
                    
                    if contact_parts:
                        contact_para = Paragraph(
                            " | ".join(contact_parts),
                            ParagraphStyle(
                                'CompanyContact',
                                parent=styles['Normal'],
                                fontSize=8,
                                textColor=colors.grey
                            )
                        )
                        company_info_data.append([contact_para])
                    
                    # Créer le tableau des infos entreprise
                    info_table = Table(
                        company_info_data,
                        colWidths=[11*cm]
                    )
                    info_table.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 0),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                        ('TOPPADDING', (0, 0), (-1, -1), 0),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ]))
                    
                    # Créer le header principal avec logo et infos
                    header_table = Table(
                        [[logo_img, info_table]],
                        colWidths=[5*cm, 11*cm],
                        rowHeights=[max(2.5*cm, logo_height + 0.5*cm)]  # Hauteur adaptative
                    )
                    
                    header_table.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                        ('LEFTPADDING', (0, 0), (0, 0), 0),
                        ('RIGHTPADDING', (0, 0), (0, 0), 0),
                        ('LEFTPADDING', (1, 0), (1, 0), 0.5*cm),
                        ('RIGHTPADDING', (1, 0), (1, 0), 0),
                        ('TOPPADDING', (0, 0), (-1, -1), 0),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                    ]))
                    
                    elements.append(header_table)
                    
                except Exception as e:
                    logger.warning(f"Erreur création header avec logo: {e}")
                    # Fallback sans logo
                    elements.append(Paragraph(tenant_name, ParagraphStyle(
                        'CompanyNameFallback', parent=styles['Heading1'], 
                        fontSize=16, textColor=primary_color, spaceAfter=6)))
                    
                    if tenant_info:
                        if tenant_info.get('address'):
                            elements.append(Paragraph(tenant_info['address'], 
                                ParagraphStyle('CompanyAddr', parent=styles['Normal'], fontSize=9, spaceAfter=3)))
                        
                        contact = []
                        if tenant_info.get('phone'):
                            contact.append(f"Tél: {tenant_info['phone']}")
                        if tenant_info.get('email'):
                            contact.append(f"Email: {tenant_info['email']}")
                        if contact:
                            elements.append(Paragraph(" | ".join(contact), 
                                ParagraphStyle('CompanyContact', parent=styles['Normal'], 
                                fontSize=8, textColor=colors.grey, spaceAfter=6)))
            else:
                # Pas de logo - layout simple
                elements.append(Paragraph(tenant_name, ParagraphStyle(
                    'CompanyNameNoLogo', parent=styles['Heading1'], 
                    fontSize=16, textColor=primary_color, spaceAfter=6)))
                
                if tenant_info:
                    if tenant_info.get('address'):
                        elements.append(Paragraph(tenant_info['address'], 
                            ParagraphStyle('CompanyAddr', parent=styles['Normal'], fontSize=9, spaceAfter=3)))
                    
                    contact = []
                    if tenant_info.get('phone'):
                        contact.append(f"Tél: {tenant_info['phone']}")
                    if tenant_info.get('email'):
                        contact.append(f"Email: {tenant_info['email']}")
                    if contact:
                        elements.append(Paragraph(" | ".join(contact), 
                            ParagraphStyle('CompanyContact', parent=styles['Normal'], 
                            fontSize=8, textColor=colors.grey, spaceAfter=6)))
            
            # Header personnalisé
            if tenant_info and tenant_info.get('header'):
                elements.append(Spacer(1, 0.3*cm))
                header_style = ParagraphStyle('CustomHeader', parent=styles['Normal'], 
                    fontSize=9, textColor=colors.grey, alignment=1, spaceAfter=6)
                elements.append(Paragraph(tenant_info['header'], header_style))
            
            elements.append(Spacer(1, 0.8*cm))
            
            # Titre reçu avec couleur primaire
            elements.append(Paragraph("REÇU DE PAIEMENT", ParagraphStyle(
                'ReceiptHeader', parent=styles['Heading2'], alignment=1,
                textColor=primary_color, fontSize=18, spaceAfter=12
            )))
            elements.append(Spacer(1, 0.5*cm))
            
            # Référence et date
            ref = payment.get('reference', payment.get('id', 'N/A')[:8])
            date_str = payment.get('created_at', '')[:10] if payment.get('created_at') else datetime.now().strftime('%Y-%m-%d')
            
            info_data = [
                ['Référence:', ref],
                ['Date:', date_str],
                ['Statut:', payment.get('status', 'N/A').upper()],
            ]
            
            info_table = Table(info_data, colWidths=[4*cm, 8*cm])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.5*cm))
            
            # Client
            elements.append(Paragraph("REÇU DE:", ParagraphStyle('ClientLabel', parent=styles['Heading3'], 
                textColor=primary_color)))
            client_name = payment.get('client_name', payment.get('client', {}).get('name', 'N/A'))
            elements.append(Paragraph(client_name, styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
            
            # Détails paiement
            elements.append(Paragraph("DÉTAILS DU PAIEMENT:", ParagraphStyle('PaymentLabel', 
                parent=styles['Heading3'], textColor=primary_color)))
            
            method_labels = {
                'cash': 'Espèces', 'mobile_money': 'Mobile Money',
                'bank_transfer': 'Virement bancaire', 'card': 'Carte bancaire'
            }
            method = method_labels.get(payment.get('method', ''), payment.get('method', 'N/A'))
            
            detail_data = [
                ['Méthode de paiement', 'Montant'],
                [method, f"{payment.get('amount', 0):,.0f} {payment.get('currency', 'XAF')}"]
            ]
            
            detail_table = Table(detail_data, colWidths=[10*cm, 6*cm])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), primary_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(detail_table)
            elements.append(Spacer(1, 0.3*cm))
            
            # Colis associés
            packages = payment.get('packages', [])
            if packages:
                elements.append(Paragraph("Colis concernés:", styles['Heading4']))
                for pkg in packages[:5]:  # Max 5 colis
                    tracking = pkg.get('tracking_number', pkg.get('tracking', 'N/A'))
                    elements.append(Paragraph(f"• {tracking}", styles['Normal']))
            
            elements.append(Spacer(1, 0.5*cm))
            
            # Total avec couleur primaire
            # Créer une couleur plus claire pour le background
            try:
                hex_color = primary_color.hexval() if hasattr(primary_color, 'hexval') else '#2563eb'
                # Convertir en RGB et éclaircir
                r = int(hex_color[1:3], 16)
                g = int(hex_color[3:5], 16)
                b = int(hex_color[5:7], 16)
                light_bg = colors.Color(min(r+50, 255)/255, min(g+50, 255)/255, min(b+50, 255)/255, alpha=0.3)
            except:
                light_bg = colors.HexColor('#dbeafe')
            
            total_data = [
                ['TOTAL PAYÉ:', f"{payment.get('amount', 0):,.0f} {payment.get('currency', 'XAF')}"]
            ]
            total_table = Table(total_data, colWidths=[10*cm, 6*cm])
            total_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 14),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('BACKGROUND', (0, 0), (-1, -1), light_bg),
                ('TEXTCOLOR', (0, 0), (-1, -1), primary_color),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(total_table)
            
            # Footer personnalisé
            if tenant_info and tenant_info.get('footer'):
                elements.append(Spacer(1, 0.5*cm))
                footer_style = ParagraphStyle('CustomFooter', parent=styles['Normal'],
                    fontSize=8, textColor=colors.grey, alignment=1)
                elements.append(Paragraph(tenant_info['footer'], footer_style))
            
            # Pied de page par défaut
            elements.append(Spacer(1, 1*cm))
            elements.append(Paragraph("Merci pour votre paiement!",
                                     ParagraphStyle('Footer', parent=styles['Normal'],
                                                   alignment=1, textColor=primary_color, fontSize=11)))
            elements.append(Paragraph(f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
                                     ParagraphStyle('FooterDate', parent=styles['Normal'],
                                                   alignment=1, fontSize=8, textColor=colors.grey)))
            
            doc.build(elements)
            
            pdf_data = buffer.getvalue()
            buffer.close()
            
            filename = f"recu_paiement_{ref}.pdf"
            
            return ExportResult(
                success=True,
                data=pdf_data,
                filename=filename,
                content_type='application/pdf'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération reçu paiement: {str(e)}")
            return ExportResult(success=False, error=str(e))
    
    def generate_pickup_receipt(self, pickup: dict, tenant_info: dict = None) -> ExportResult:
        """
        Génère un reçu de retrait en PDF
        
        Args:
            pickup: Données du retrait
            tenant_info: Infos du tenant (logo, header, footer, primary_color)
        
        Returns:
            ExportResult avec le PDF
        """
        if not REPORTLAB_AVAILABLE:
            return ExportResult(success=False, error="reportlab non installé")
        
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4,
                                   rightMargin=2*cm, leftMargin=2*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)
            
            elements = []
            styles = self.styles
            primary_color = self._get_primary_color(tenant_info)
            
            # Logo
            self._add_logo(elements, tenant_info)
            
            # Header personnalisé
            self._add_custom_header(elements, tenant_info)
            
            # Style titre avec couleur primaire
            title_style = ParagraphStyle(
                'PickupTitle',
                parent=styles['Heading1'],
                fontSize=20,
                spaceAfter=20,
                textColor=primary_color,
                alignment=1
            )
            
            # En-tête entreprise
            tenant_name = tenant_info.get('name', self.tenant_name) if tenant_info else self.tenant_name
            elements.append(Paragraph(tenant_name, title_style))
            
            if tenant_info:
                info_style = ParagraphStyle('InfoCenter', parent=styles['Normal'], alignment=1)
                if tenant_info.get('address'):
                    elements.append(Paragraph(tenant_info['address'], info_style))
                contact = []
                if tenant_info.get('phone'):
                    contact.append(f"Tél: {tenant_info['phone']}")
                if tenant_info.get('email'):
                    contact.append(f"Email: {tenant_info['email']}")
                if contact:
                    elements.append(Paragraph(" | ".join(contact), info_style))
            
            elements.append(Spacer(1, 1*cm))
            
            # Titre reçu avec couleur primaire
            elements.append(Paragraph("REÇU DE RETRAIT", ParagraphStyle(
                'PickupHeader', parent=styles['Heading2'], alignment=1,
                textColor=primary_color
            )))
            elements.append(Spacer(1, 0.5*cm))
            
            # Infos retrait
            pickup_date = pickup.get('picked_up_at', pickup.get('pickup_date', ''))[:10] if pickup.get('picked_up_at') or pickup.get('pickup_date') else datetime.now().strftime('%Y-%m-%d')
            
            info_data = [
                ['N° Retrait:', pickup.get('id', 'N/A')[:8]],
                ['Date:', pickup_date],
                ['Tracking colis:', pickup.get('package_tracking', pickup.get('tracking_number', 'N/A'))],
            ]
            
            info_table = Table(info_data, colWidths=[4*cm, 8*cm])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.5*cm))
            
            # Personne qui retire
            elements.append(Paragraph("RETIRÉ PAR:", styles['Heading3']))
            picker_name = pickup.get('picker_name', pickup.get('picked_by_name', 'N/A'))
            picker_phone = pickup.get('picker_phone', pickup.get('picked_by_phone', ''))
            picker_id = pickup.get('picker_id_number', pickup.get('id_number', ''))
            
            elements.append(Paragraph(f"Nom: {picker_name}", styles['Normal']))
            if picker_phone:
                elements.append(Paragraph(f"Téléphone: {picker_phone}", styles['Normal']))
            if picker_id:
                elements.append(Paragraph(f"N° Pièce d'identité: {picker_id}", styles['Normal']))
            
            elements.append(Spacer(1, 0.5*cm))
            
            # Détails colis
            elements.append(Paragraph("DÉTAILS DU COLIS:", styles['Heading3']))
            
            package = pickup.get('package', {})
            detail_data = [
                ['Description', package.get('description', 'N/A')[:50]],
                ['Poids', f"{package.get('weight', 'N/A')} kg"],
                ['Quantité', str(package.get('quantity', 1))],
            ]
            
            detail_table = Table(detail_data, colWidths=[4*cm, 10*cm])
            detail_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(detail_table)
            elements.append(Spacer(1, 0.5*cm))
            
            # Paiement collecté
            amount_collected = pickup.get('payment_collected', pickup.get('amount_collected', 0))
            if amount_collected and amount_collected > 0:
                elements.append(Paragraph("PAIEMENT COLLECTÉ:", styles['Heading3']))
                payment_data = [
                    ['Montant:', f"{amount_collected:,.0f} XAF"]
                ]
                payment_table = Table(payment_data, colWidths=[4*cm, 10*cm])
                payment_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 12),
                    ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#059669')),
                ]))
                elements.append(payment_table)
                elements.append(Spacer(1, 0.5*cm))
            
            # Signature
            elements.append(Spacer(1, 1*cm))
            elements.append(Paragraph("SIGNATURE:", styles['Heading4']))
            
            # Zone signature (cadre vide)
            sig_table = Table([['Signature du retrait']], colWidths=[8*cm])
            sig_table.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 30),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(sig_table)
            
            # Footer personnalisé
            self._add_custom_footer(elements, tenant_info)
            
            # Pied de page par défaut
            elements.append(Spacer(1, 1*cm))
            elements.append(Paragraph("Ce document atteste du retrait du colis mentionné ci-dessus.",
                                     ParagraphStyle('Footer', parent=styles['Normal'],
                                                   alignment=1, fontSize=9, textColor=colors.grey)))
            elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
                                     ParagraphStyle('FooterDate', parent=styles['Normal'],
                                                   alignment=1, fontSize=8, textColor=colors.grey)))
            
            doc.build(elements)
            
            pdf_data = buffer.getvalue()
            buffer.close()
            
            tracking = pickup.get('package_tracking', pickup.get('tracking_number', 'unknown'))
            filename = f"recu_retrait_{tracking}.pdf"
            
            return ExportResult(
                success=True,
                data=pdf_data,
                filename=filename,
                content_type='application/pdf'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération reçu retrait: {str(e)}")
            return ExportResult(success=False, error=str(e))
    
    def generate_statistics_report(self, stats: dict, tenant_info: dict = None) -> ExportResult:
        """
        Génère un rapport statistiques en PDF
        
        Args:
            stats: Données statistiques (depuis /finance/stats)
            tenant_info: Infos du tenant (logo, header, footer, primary_color)
        
        Returns:
            ExportResult avec le PDF
        """
        if not REPORTLAB_AVAILABLE:
            return ExportResult(success=False, error="reportlab non installé")
        
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4,
                                   rightMargin=2*cm, leftMargin=2*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)
            
            elements = []
            styles = self.styles
            primary_color = self._get_primary_color(tenant_info)
            
            # Logo
            self._add_logo(elements, tenant_info)
            
            # Header personnalisé
            self._add_custom_header(elements, tenant_info)
            
            # Style titre avec couleur primaire
            title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontSize=22,
                spaceAfter=20,
                textColor=primary_color,
                alignment=1
            )
            
            # En-tête
            tenant_name = tenant_info.get('name', self.tenant_name) if tenant_info else self.tenant_name
            elements.append(Paragraph(tenant_name, title_style))
            elements.append(Spacer(1, 0.5*cm))
            
            # Titre rapport
            period = stats.get('period', {})
            period_str = f"{period.get('start', '')} au {period.get('end', '')}"
            elements.append(Paragraph("RAPPORT STATISTIQUES", styles['Heading2']))
            elements.append(Paragraph(f"Période: {period_str}", ParagraphStyle(
                'Period', parent=styles['Normal'], alignment=1, textColor=colors.grey
            )))
            elements.append(Spacer(1, 1*cm))
            
            # KPIs principaux
            elements.append(Paragraph("INDICATEURS CLÉS", styles['Heading3']))
            elements.append(Spacer(1, 0.3*cm))
            
            revenue = stats.get('revenue', {})
            packages = stats.get('packages', {})
            clients = stats.get('clients', {})
            delivery = stats.get('delivery', {})
            
            kpi_data = [
                ['Indicateur', 'Valeur', 'Évolution'],
                ['Chiffre d\'affaires', f"{revenue.get('total', 0):,.0f} XAF", self._format_change(revenue.get('total', 0), revenue.get('previous', 0))],
                ['Colis traités', str(packages.get('count', 0)), self._format_change(packages.get('count', 0), packages.get('count_previous', 0))],
                ['Nouveaux clients', str(clients.get('new', 0)), self._format_change(clients.get('new', 0), clients.get('new_previous', 0))],
                ['Taux de livraison', f"{delivery.get('rate', 0)}%", self._format_change(delivery.get('rate', 0), delivery.get('rate_previous', 0))],
                ['Impayés', f"{packages.get('unpaid_amount', 0):,.0f} XAF", '-'],
            ]
            
            kpi_table = Table(kpi_data, colWidths=[6*cm, 5*cm, 4*cm])
            kpi_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(kpi_table)
            elements.append(Spacer(1, 1*cm))
            
            # Répartition par transport
            by_transport = packages.get('by_transport', {})
            if by_transport:
                elements.append(Paragraph("RÉPARTITION PAR MODE DE TRANSPORT", styles['Heading3']))
                elements.append(Spacer(1, 0.3*cm))
                
                transport_labels = {'sea': 'Maritime', 'air_normal': 'Aérien Normal', 'air_express': 'Aérien Express'}
                transport_data = [['Mode', 'Colis', 'Revenus']]
                
                for mode, data in by_transport.items():
                    transport_data.append([
                        transport_labels.get(mode, mode),
                        str(data.get('count', 0)),
                        f"{data.get('revenue', 0):,.0f} XAF"
                    ])
                
                transport_table = Table(transport_data, colWidths=[5*cm, 4*cm, 5*cm])
                transport_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                ]))
                elements.append(transport_table)
                elements.append(Spacer(1, 1*cm))
            
            # Top clients
            top_clients = stats.get('top_clients', [])
            if top_clients:
                elements.append(Paragraph("TOP 10 CLIENTS", styles['Heading3']))
                elements.append(Spacer(1, 0.3*cm))
                
                clients_data = [['#', 'Client', 'Colis', 'CA Total']]
                for i, client in enumerate(top_clients[:10], 1):
                    clients_data.append([
                        str(i),
                        client.get('name', 'N/A')[:25],
                        str(client.get('packages', 0)),
                        f"{client.get('revenue', 0):,.0f} XAF"
                    ])
                
                clients_table = Table(clients_data, colWidths=[1*cm, 7*cm, 2.5*cm, 4*cm])
                clients_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                ]))
                elements.append(clients_table)
            
            # Footer personnalisé
            self._add_custom_footer(elements, tenant_info)
            
            # Pied de page par défaut
            elements.append(Spacer(1, 2*cm))
            elements.append(Paragraph(f"Rapport généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
                                     ParagraphStyle('FooterDate', parent=styles['Normal'],
                                                   alignment=1, fontSize=8, textColor=colors.grey)))
            
            doc.build(elements)
            
            pdf_data = buffer.getvalue()
            buffer.close()
            
            filename = f"rapport_statistiques_{datetime.now().strftime('%Y%m%d')}.pdf"
            
            return ExportResult(
                success=True,
                data=pdf_data,
                filename=filename,
                content_type='application/pdf'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération rapport stats: {str(e)}")
            return ExportResult(success=False, error=str(e))
    
    # ==================== TICKETS (Format 80mm) ====================
    
    def generate_payment_ticket(self, payment: dict, tenant_info: dict = None) -> ExportResult:
        """
        Génère un ticket de paiement (format 80mm pour imprimante thermique)
        
        Args:
            payment: Données du paiement
            tenant_info: Infos du tenant
        
        Returns:
            ExportResult avec le PDF ticket
        """
        if not REPORTLAB_AVAILABLE:
            return ExportResult(success=False, error="reportlab non installé")
        
        try:
            buffer = io.BytesIO()
            # Format ticket: 80mm de large, hauteur variable
            page_width = 80 * mm
            page_height = 250 * mm  # Hauteur max, sera coupée
            
            c = canvas.Canvas(buffer, pagesize=(page_width, page_height))
            primary_color = self._get_primary_color(tenant_info)
            
            y = page_height - 10*mm
            center_x = page_width / 2
            
            # === LOGO (si disponible) ===
            if tenant_info and tenant_info.get('logo') and tenant_info.get('logo').startswith('data:image'):
                try:
                    import base64
                    logo_data = tenant_info['logo']
                    if ',' in logo_data:
                        logo_base64 = logo_data.split(',')[1]
                    else:
                        logo_base64 = logo_data
                    logo_bytes = base64.b64decode(logo_base64)
                    logo_buffer = io.BytesIO(logo_bytes)
                    
                    # Dessiner le logo centré
                    from reportlab.lib.utils import ImageReader
                    img_reader = ImageReader(logo_buffer)
                    img_width = 25*mm
                    img_height = 12*mm
                    c.drawImage(img_reader, center_x - img_width/2, y - img_height, 
                               width=img_width, height=img_height, preserveAspectRatio=True, mask='auto')
                    y -= img_height + 3*mm
                except Exception as e:
                    logger.warning(f"Erreur logo ticket: {e}")
            
            # === EN-TÊTE ===
            tenant_name = tenant_info.get('name', self.tenant_name) if tenant_info else self.tenant_name
            c.setFillColor(primary_color)
            c.setFont("Helvetica-Bold", 12)
            c.drawCentredString(center_x, y, tenant_name)
            y -= 5*mm
            
            c.setFillColor(colors.black)
            if tenant_info:
                c.setFont("Helvetica", 7)
                if tenant_info.get('address'):
                    c.drawCentredString(center_x, y, tenant_info['address'][:40])
                    y -= 3.5*mm
                if tenant_info.get('phone'):
                    c.drawCentredString(center_x, y, f"Tél: {tenant_info['phone']}")
                    y -= 3.5*mm
            
            # Ligne séparatrice avec couleur primaire
            y -= 2*mm
            c.setStrokeColor(primary_color)
            c.setLineWidth(0.5)
            c.line(5*mm, y, page_width - 5*mm, y)
            c.setStrokeColor(colors.black)
            y -= 5*mm
            
            # === TITRE ===
            c.setFillColor(primary_color)
            c.setFont("Helvetica-Bold", 11)
            c.drawCentredString(center_x, y, "REÇU DE PAIEMENT")
            y -= 6*mm
            
            # === INFOS PAIEMENT ===
            c.setFillColor(colors.black)
            c.setFont("Helvetica", 8)
            ref = payment.get('reference', payment.get('id', 'N/A')[:8])
            date_str = payment.get('created_at', '')[:16] if payment.get('created_at') else datetime.now().strftime('%Y-%m-%d %H:%M')
            
            c.drawString(5*mm, y, f"Réf: {ref}")
            y -= 4*mm
            c.drawString(5*mm, y, f"Date: {date_str}")
            y -= 4*mm
            
            # Client
            client_name = payment.get('client_name', payment.get('client', {}).get('name', 'N/A'))
            c.drawString(5*mm, y, f"Client: {client_name[:25]}")
            y -= 5*mm
            
            # Ligne séparatrice
            c.setDash(1, 2)
            c.line(5*mm, y, page_width - 5*mm, y)
            c.setDash()
            y -= 5*mm
            
            # === DÉTAILS ===
            method_labels = {
                'cash': 'Espèces', 'mobile_money': 'Mobile Money',
                'bank_transfer': 'Virement', 'card': 'Carte'
            }
            method = method_labels.get(payment.get('method', ''), payment.get('method', 'N/A'))
            
            c.setFont("Helvetica", 8)
            c.drawString(5*mm, y, f"Mode: {method}")
            y -= 4*mm
            
            # Colis associés
            packages = payment.get('packages', [])
            if packages:
                c.drawString(5*mm, y, "Colis:")
                y -= 4*mm
                for pkg in packages[:3]:
                    tracking = pkg.get('tracking_number', pkg.get('tracking', 'N/A'))
                    c.drawString(8*mm, y, f"• {tracking}")
                    y -= 3.5*mm
                if len(packages) > 3:
                    c.drawString(8*mm, y, f"  +{len(packages)-3} autre(s)")
                    y -= 3.5*mm
            
            y -= 2*mm
            
            # === MONTANT ===
            c.setStrokeColor(primary_color)
            c.setLineWidth(0.5)
            c.line(5*mm, y, page_width - 5*mm, y)
            c.setStrokeColor(colors.black)
            y -= 6*mm
            
            amount = payment.get('amount', 0)
            currency = payment.get('currency', 'XAF')
            
            c.setFillColor(primary_color)
            c.setFont("Helvetica-Bold", 12)
            c.drawCentredString(center_x, y, f"TOTAL: {amount:,.0f} {currency}")
            y -= 8*mm
            
            # Statut
            status = payment.get('status', 'completed')
            status_text = "✓ PAYÉ" if status == 'completed' else status.upper()
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(center_x, y, status_text)
            y -= 6*mm
            
            # === FOOTER ===
            c.setFillColor(colors.black)
            c.setDash(1, 2)
            c.line(5*mm, y, page_width - 5*mm, y)
            c.setDash()
            y -= 5*mm
            
            # Footer personnalisé
            if tenant_info and tenant_info.get('footer'):
                c.setFont("Helvetica", 6)
                footer_lines = tenant_info['footer'][:80].split('\n')
                for line in footer_lines[:2]:  # Max 2 lignes
                    c.drawCentredString(center_x, y, line.strip())
                    y -= 3*mm
            
            c.setFont("Helvetica", 7)
            c.drawCentredString(center_x, y, "Merci pour votre paiement!")
            y -= 3.5*mm
            c.drawCentredString(center_x, y, datetime.now().strftime('%d/%m/%Y %H:%M'))
            
            c.save()
            
            pdf_data = buffer.getvalue()
            buffer.close()
            
            ref = payment.get('reference', payment.get('id', 'N/A')[:8])
            filename = f"ticket_paiement_{ref}.pdf"
            
            return ExportResult(
                success=True,
                data=pdf_data,
                filename=filename,
                content_type='application/pdf'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération ticket paiement: {str(e)}")
            return ExportResult(success=False, error=str(e))
    
    def generate_pickup_ticket(self, pickup: dict, tenant_info: dict = None) -> ExportResult:
        """
        Génère un ticket de retrait (format 80mm pour imprimante thermique)
        
        Args:
            pickup: Données du retrait
            tenant_info: Infos du tenant
        
        Returns:
            ExportResult avec le PDF ticket
        """
        if not REPORTLAB_AVAILABLE:
            return ExportResult(success=False, error="reportlab non installé")
        
        try:
            buffer = io.BytesIO()
            page_width = 80 * mm
            page_height = 220 * mm
            
            c = canvas.Canvas(buffer, pagesize=(page_width, page_height))
            
            y = page_height - 10*mm
            center_x = page_width / 2
            
            # === EN-TÊTE ===
            tenant_name = tenant_info.get('name', self.tenant_name) if tenant_info else self.tenant_name
            c.setFont("Helvetica-Bold", 12)
            c.drawCentredString(center_x, y, tenant_name)
            y -= 5*mm
            
            if tenant_info:
                c.setFont("Helvetica", 7)
                if tenant_info.get('address'):
                    c.drawCentredString(center_x, y, tenant_info['address'][:40])
                    y -= 3.5*mm
                if tenant_info.get('phone'):
                    c.drawCentredString(center_x, y, f"Tél: {tenant_info['phone']}")
                    y -= 3.5*mm
            
            # Ligne séparatrice
            y -= 2*mm
            c.setDash(1, 2)
            c.line(5*mm, y, page_width - 5*mm, y)
            c.setDash()
            y -= 5*mm
            
            # === TITRE ===
            c.setFont("Helvetica-Bold", 11)
            c.drawCentredString(center_x, y, "REÇU DE RETRAIT")
            y -= 6*mm
            
            # === INFOS RETRAIT ===
            c.setFont("Helvetica", 8)
            pickup_id = pickup.get('id', 'N/A')[:8]
            pickup_date = pickup.get('picked_up_at', pickup.get('pickup_date', ''))[:16] if pickup.get('picked_up_at') or pickup.get('pickup_date') else datetime.now().strftime('%Y-%m-%d %H:%M')
            
            c.drawString(5*mm, y, f"N°: {pickup_id}")
            y -= 4*mm
            c.drawString(5*mm, y, f"Date: {pickup_date}")
            y -= 5*mm
            
            # Ligne séparatrice
            c.setDash(1, 2)
            c.line(5*mm, y, page_width - 5*mm, y)
            c.setDash()
            y -= 5*mm
            
            # === COLIS ===
            c.setFont("Helvetica-Bold", 9)
            c.drawString(5*mm, y, "COLIS")
            y -= 4*mm
            
            tracking = pickup.get('package_tracking', pickup.get('tracking_number', 'N/A'))
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(center_x, y, tracking)
            y -= 5*mm
            
            package = pickup.get('package', {})
            c.setFont("Helvetica", 8)
            if package.get('description'):
                desc = package['description'][:35]
                c.drawString(5*mm, y, desc)
                y -= 4*mm
            
            c.drawString(5*mm, y, f"Poids: {package.get('weight', 'N/A')} kg")
            c.drawString(40*mm, y, f"Qté: {package.get('quantity', 1)}")
            y -= 5*mm
            
            # Ligne séparatrice
            c.setDash(1, 2)
            c.line(5*mm, y, page_width - 5*mm, y)
            c.setDash()
            y -= 5*mm
            
            # === PERSONNE QUI RETIRE ===
            c.setFont("Helvetica-Bold", 9)
            c.drawString(5*mm, y, "RETIRÉ PAR")
            y -= 4*mm
            
            picker_name = pickup.get('picker_name', pickup.get('picked_by_name', 'N/A'))
            picker_phone = pickup.get('picker_phone', pickup.get('picked_by_phone', ''))
            picker_id = pickup.get('picker_id_number', pickup.get('id_number', ''))
            
            c.setFont("Helvetica", 8)
            c.drawString(5*mm, y, f"Nom: {picker_name[:25]}")
            y -= 4*mm
            if picker_phone:
                c.drawString(5*mm, y, f"Tél: {picker_phone}")
                y -= 4*mm
            if picker_id:
                c.drawString(5*mm, y, f"CNI: {picker_id}")
                y -= 4*mm
            
            y -= 2*mm
            
            # === PAIEMENT COLLECTÉ ===
            amount_collected = pickup.get('payment_collected', pickup.get('amount_collected', 0))
            if amount_collected and amount_collected > 0:
                c.setDash(1, 2)
                c.line(5*mm, y, page_width - 5*mm, y)
                c.setDash()
                y -= 5*mm
                
                c.setFont("Helvetica-Bold", 9)
                c.drawString(5*mm, y, "PAIEMENT COLLECTÉ")
                y -= 5*mm
                
                c.setFont("Helvetica-Bold", 11)
                c.drawCentredString(center_x, y, f"{amount_collected:,.0f} XAF")
                y -= 5*mm
            
            # === SIGNATURE ===
            y -= 3*mm
            c.setDash(1, 2)
            c.line(5*mm, y, page_width - 5*mm, y)
            c.setDash()
            y -= 5*mm
            
            c.setFont("Helvetica", 7)
            c.drawString(5*mm, y, "Signature:")
            y -= 12*mm
            c.line(5*mm, y, page_width - 5*mm, y)
            y -= 5*mm
            
            # === FOOTER ===
            c.setFont("Helvetica", 7)
            c.drawCentredString(center_x, y, "Colis retiré avec succès")
            y -= 3.5*mm
            c.drawCentredString(center_x, y, datetime.now().strftime('%d/%m/%Y %H:%M'))
            
            c.save()
            
            pdf_data = buffer.getvalue()
            buffer.close()
            
            filename = f"ticket_retrait_{tracking}.pdf"
            
            return ExportResult(
                success=True,
                data=pdf_data,
                filename=filename,
                content_type='application/pdf'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération ticket retrait: {str(e)}")
            return ExportResult(success=False, error=str(e))
    
    def _format_change(self, current, previous):
        """Formate le changement en pourcentage"""
        if not previous or previous == 0:
            return '-'
        change = ((current - previous) / previous) * 100
        sign = '+' if change >= 0 else ''
        return f"{sign}{change:.1f}%"


class ExcelGenerator:
    """Générateur de fichiers Excel"""
    
    def __init__(self):
        pass
    
    def generate_packages_excel(self, packages: List[dict], title: str = "Liste des colis") -> ExportResult:
        """
        Génère un export Excel des colis
        
        Args:
            packages: Liste des colis (dicts)
            title: Titre du rapport
        
        Returns:
            ExportResult avec le fichier Excel
        """
        if not OPENPYXL_AVAILABLE:
            return ExportResult(success=False, error="openpyxl non installé")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Colis"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # En-têtes
            headers = [
                "Tracking", "Client", "Description", "Mode Transport",
                "Poids (kg)", "Quantité", "Statut", "Montant", "Payé",
                "Destination", "Date création"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Données
            for row, pkg in enumerate(packages, 2):
                ws.cell(row=row, column=1, value=pkg.get('tracking_number', ''))
                
                client = pkg.get('client', {})
                ws.cell(row=row, column=2, value=client.get('name', '') if isinstance(client, dict) else '')
                
                ws.cell(row=row, column=3, value=pkg.get('description', '')[:50])
                ws.cell(row=row, column=4, value=pkg.get('transport_mode', ''))
                ws.cell(row=row, column=5, value=pkg.get('weight', ''))
                ws.cell(row=row, column=6, value=pkg.get('quantity', 1))
                ws.cell(row=row, column=7, value=pkg.get('status', ''))
                ws.cell(row=row, column=8, value=pkg.get('amount', 0))
                ws.cell(row=row, column=9, value=pkg.get('paid_amount', 0))
                
                dest = pkg.get('destination', {})
                ws.cell(row=row, column=10, value=dest.get('city', '') if isinstance(dest, dict) else '')
                
                ws.cell(row=row, column=11, value=pkg.get('created_at', '')[:10] if pkg.get('created_at') else '')
                
                # Bordures
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).border = thin_border
            
            # Ajuster largeur colonnes
            column_widths = [15, 20, 30, 12, 10, 8, 12, 12, 12, 15, 12]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # Sauvegarder
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"colis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return ExportResult(
                success=True,
                data=buffer.getvalue(),
                filename=filename,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération Excel colis: {str(e)}")
            return ExportResult(success=False, error=str(e))
    
    def generate_invoices_excel(self, invoices: List[dict]) -> ExportResult:
        """Génère un export Excel des factures"""
        if not OPENPYXL_AVAILABLE:
            return ExportResult(success=False, error="openpyxl non installé")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Factures"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
            
            headers = ["N° Facture", "Client", "Description", "Montant", "Devise", 
                      "Statut", "Date émission", "Date échéance", "Date paiement"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            for row, inv in enumerate(invoices, 2):
                ws.cell(row=row, column=1, value=inv.get('invoice_number', ''))
                ws.cell(row=row, column=2, value=inv.get('client_name', ''))
                ws.cell(row=row, column=3, value=inv.get('description', '')[:50])
                ws.cell(row=row, column=4, value=inv.get('amount', 0))
                ws.cell(row=row, column=5, value=inv.get('currency', 'XAF'))
                ws.cell(row=row, column=6, value=inv.get('status', ''))
                ws.cell(row=row, column=7, value=inv.get('issue_date', ''))
                ws.cell(row=row, column=8, value=inv.get('due_date', ''))
                ws.cell(row=row, column=9, value=inv.get('paid_at', ''))
            
            # Ajuster colonnes
            for i in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 15
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"factures_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return ExportResult(
                success=True,
                data=buffer.getvalue(),
                filename=filename,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération Excel factures: {str(e)}")
            return ExportResult(success=False, error=str(e))
    
    def generate_departures_excel(self, departures: List[dict]) -> ExportResult:
        """Génère un export Excel des départs"""
        if not OPENPYXL_AVAILABLE:
            return ExportResult(success=False, error="openpyxl non installé")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Départs"
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
            
            headers = ["Origine", "Destination", "Mode", "Date départ", 
                      "Durée (j)", "Arrivée estimée", "Statut", "Nb colis", "Référence"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            for row, dep in enumerate(departures, 2):
                origin = f"{dep.get('origin_city', '')} {dep.get('origin_country', '')}".strip()
                ws.cell(row=row, column=1, value=origin)
                ws.cell(row=row, column=2, value=dep.get('dest_country', ''))
                ws.cell(row=row, column=3, value=dep.get('transport_mode', ''))
                ws.cell(row=row, column=4, value=dep.get('departure_date', ''))
                ws.cell(row=row, column=5, value=dep.get('estimated_duration', ''))
                ws.cell(row=row, column=6, value=dep.get('estimated_arrival', ''))
                ws.cell(row=row, column=7, value=dep.get('status', ''))
                ws.cell(row=row, column=8, value=dep.get('packages_count', 0))
                ws.cell(row=row, column=9, value=dep.get('reference', ''))
            
            for i in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 15
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"departs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return ExportResult(
                success=True,
                data=buffer.getvalue(),
                filename=filename,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération Excel départs: {str(e)}")
            return ExportResult(success=False, error=str(e))
    
    def generate_payments_excel(self, payments: List[dict]) -> ExportResult:
        """Génère un export Excel des paiements"""
        if not OPENPYXL_AVAILABLE:
            return ExportResult(success=False, error="openpyxl non installé")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Paiements"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # En-têtes
            headers = [
                "Référence", "Client", "Téléphone", "Montant", "Devise",
                "Méthode", "Statut", "Date", "Colis", "Notes"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Données
            for row, payment in enumerate(payments, 2):
                ws.cell(row=row, column=1, value=payment.get('reference', ''))
                ws.cell(row=row, column=2, value=payment.get('client_name', ''))
                ws.cell(row=row, column=3, value=payment.get('client_phone', ''))
                ws.cell(row=row, column=4, value=payment.get('amount', 0))
                ws.cell(row=row, column=5, value=payment.get('currency', 'XAF'))
                
                # Méthode de paiement avec labels lisibles
                method_labels = {
                    'cash': 'Espèces',
                    'mobile_money': 'Mobile Money',
                    'bank_transfer': 'Virement',
                    'card': 'Carte bancaire'
                }
                method = method_labels.get(payment.get('method', ''), payment.get('method', ''))
                ws.cell(row=row, column=6, value=method)
                
                # Statut avec labels lisibles
                status_labels = {
                    'pending': 'En attente',
                    'completed': 'Confirmé',
                    'failed': 'Échoué',
                    'cancelled': 'Annulé'
                }
                status = status_labels.get(payment.get('status', ''), payment.get('status', ''))
                ws.cell(row=row, column=7, value=status)
                
                ws.cell(row=row, column=8, value=payment.get('created_at', ''))
                
                # Colis associés (liste des tracking numbers)
                packages = payment.get('packages', [])
                if packages:
                    tracking_list = [pkg.get('tracking_number', pkg.get('tracking', '')) for pkg in packages[:3]]
                    if len(packages) > 3:
                        tracking_list.append(f"... +{len(packages)-3}")
                    ws.cell(row=row, column=9, value=", ".join(tracking_list))
                else:
                    ws.cell(row=row, column=9, value="")
                
                ws.cell(row=row, column=10, value=payment.get('notes', ''))
                
                # Appliquer les bordures
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).border = thin_border
            
            # Ajuster les largeurs de colonnes
            column_widths = [15, 20, 15, 12, 8, 15, 12, 15, 25, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # Sauvegarder
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"paiements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return ExportResult(
                success=True,
                data=buffer.getvalue(),
                filename=filename,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            logger.error(f"Erreur génération Excel paiements: {str(e)}")
            return ExportResult(success=False, error=str(e))
